import cv2
import glob
import numpy as np
import time
from pyzbar.pyzbar import decode
from PIL import Image
from openpyxl import Workbook
from openpyxl import load_workbook

def barcodeReader(image): # Decode the barcode image
    detectedBarcodes = decode(image)

    if not detectedBarcodes: # If not detected then print the message
        print("Error with this image, please try again...")
    else:
        for barcode in detectedBarcodes:
            if barcode.data != "":
                data = barcode.data.decode('utf-8')
                return data.split(",")

def reorder(myPoints):
    myPoints = myPoints.reshape((4, 2)) # REMOVE EXTRA BRACKET
    myPointsNew = np.zeros((4, 1, 2), np.int32) # NEW MATRIX WITH ARRANGED POINTS
    add = myPoints.sum(1)
    myPointsNew[0] = myPoints[np.argmin(add)]  #[0,0]
    myPointsNew[3] =myPoints[np.argmax(add)]   #[w,h]
    diff = np.diff(myPoints, axis=1)
    myPointsNew[1] =myPoints[np.argmin(diff)]  #[w,0]
    myPointsNew[2] = myPoints[np.argmax(diff)] #[h,0]
    return myPointsNew

def rectContour(contours):
    rectCon = []
    for i in contours:
        area = cv2.contourArea(i)
        if area > 50:
            peri = cv2.arcLength(i, True)
            approx = cv2.approxPolyDP(i, 0.02 * peri, True)
            if len(approx) == 4:
                rectCon.append(i)
    rectCon = sorted(rectCon, key=cv2.contourArea,reverse=True)
    return rectCon

def getCornerPoints(cont):
    peri = cv2.arcLength(cont, True) # LENGTH OF CONTOUR
    approx = cv2.approxPolyDP(cont, 0.02 * peri, True) # APPROXIMATE THE POLY TO GET CORNER POINTS
    return approx

def splitBoxes(img, questions, choices): # CREATE A SPECIFIC BOX FOR EACH CHOICES
    rows = np.vsplit(img, questions) # SPLIT OMR IMAGE VERTICALLY BASED ON QUESTIONS
    boxes={}
    for i,r in enumerate(rows, start = 1):
        cols= np.hsplit(r,choices) # SPLIT OMR IMAGE HORIZONTALLY BASED ON CHOICES
        boxes[i] = []
        for box in cols:
            boxes[i].append(box) # APPEND CHOICES INTO EACH QUESTIONS
    return boxes

start = time.time() # USED AS START TIME
threasholdValue = 1000 # VALUE DETERMINE APPROXIMATE CIRCLE COLOURED VALUE
pathImage = "Datas/*.jpg" # OMR SHEET DATA PATH
ans= {1: 1, 2: 3, 3: 2, 4: 3, 5: 2, 6: 2, 7: 2, 8: 1, 9: 3, 10: 4}  # ANSWER USED FOR VALIDATE
path = glob.glob(pathImage)

for imageData in path: # ITTERATE OVER EACH IMAGE IN PATH
    img = cv2.imread(imageData)
    barcodeData = barcodeReader(Image.open(imageData)) # READ BARCODE DATAS
    questions = int(barcodeData[0])
    choices = int(barcodeData[1])
    regno = barcodeData[2]
    answers = []
    category = "Not_Categorized" if len(barcodeData) == 3 else barcodeData[3]
    filename = "Results/" + category + ".xlsx" # EXCEL SHEET FILE PATH

    try:
        wb = load_workbook(filename)
        ws = wb.worksheets[0]  # SELECT FIRST WORKSHEET
    except FileNotFoundError:
        headers_row = ['Name', 'Marks scored']
        wb = Workbook()
        ws = wb.active

        for q in range(questions):
            headers_row.append("Q" + repr(q + 1)) # MAKING HEADER ROW

        ws.append(headers_row)

    heightImg = questions * 100
    widthImg = choices * 100
    img = cv2.resize(img, (widthImg, heightImg))  # RESIZE IMAGE
    imgGray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)  # CONVERT IMAGE TO GRAY SCALE
    imgBlur = cv2.GaussianBlur(imgGray, (5, 5), 1)  # ADD GAUSSIAN BLUR
    imgCanny = cv2.Canny(imgBlur, 10, 70)  # APPLY CANNY

    try:
        contours, hierarchy = cv2.findContours(imgCanny, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_NONE)  # FIND ALL CONTOURS
        rectCon = rectContour(contours)  # FILTER FOR RECTANGLE CONTOURS
        biggestPoints = getCornerPoints(rectCon[0])  # GET CORNER POINTS OF THE BIGGEST RECTANGLE

        if biggestPoints.size != 0:  # BIGGEST RECTANGLE WARPING
            biggestPoints = reorder(biggestPoints)  # REORDER FOR WARPING
            pts1 = np.float32(biggestPoints)  # PREPARE POINTS FOR WARP
            pts2 = np.float32([[0, 0], [widthImg, 0], [0, heightImg], [widthImg, heightImg]])  # PREPARE POINTS FOR WARP
            matrix = cv2.getPerspectiveTransform(pts1, pts2)  # GET TRANSFORMATION MATRIX
            imgWarpColored = cv2.warpPerspective(img, matrix, (widthImg, heightImg))  # APPLY WARP PERSPECTIVE

            # APPLY THRESHOLD
            imgWarpGray = cv2.cvtColor(imgWarpColored, cv2.COLOR_BGR2GRAY)  # CONVERT TO GRAYSCALE
            # imgThresh =cv2.adaptiveThreshold(imgWarpGray, 225, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 11, 2)
            imgThresh = cv2.threshold(imgWarpGray, 170, 255, cv2.THRESH_BINARY_INV)[1]  # APPLY THRESHOLD AND INVERSE
            boxes = splitBoxes(imgThresh, questions, choices)  # GET INDIVIDUAL BOXES
            choiceSelected = 0
            answersMarked = {}

            for question, choice in boxes.items():
                choiceSelected = -1

                for i,c in enumerate(choice, start = 1) :
                    markedArea = (np.sum(c == 255))  # FIND WHITE PIXELS INSIDE BOXES

                    if markedArea > threasholdValue :  # COMPARE BLACK PIXELS WITH THRESHOLD VALUES
                        choiceSelected = i if choiceSelected == -1 else -2

                answersMarked.update({question: choiceSelected}) # MAPPING QUESTIONS AND ANSWERS

                if choiceSelected > 0 :
                    answers.append(choiceSelected)
                elif choiceSelected == -1 :
                    answers.append("Not answered")
                elif choiceSelected == -2 :
                    answers.append("More than one choice selected")

            print("USER ANSWERS", answersMarked)
            correctAnswers = 0

            for x, y in answersMarked.items():
                if ans[x] == y:  # FIND CORRECT ANSWER
                    correctAnswers += 1

            score = (correctAnswers / questions) * 100  # FINAL GRADE
            print("CORRECT ANSWERS : ", correctAnswers)
            print("SCORE :", score)
            ws.append([regno, correctAnswers] + answers) # APPEND STUDENT DATAS INTO EXCEL SHEET
            wb.save(filename)

    except Exception as e:
        print(str(e))

print ("RUNNING TIME :", time.time()-start, "SECONDS")
print("TOTAL IMAGE PROCESSED", len(path))